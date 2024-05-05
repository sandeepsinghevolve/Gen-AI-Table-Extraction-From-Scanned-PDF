import boto3
import json
import openpyxl

# Configure AWS credentials (replace with your own)
session = boto3.Session(aws_access_key_id='YOUR_ACCESS_KEY_ID',
                        aws_secret_access_key='YOUR_SECRET_ACCESS_KEY')
textract = session.client('textract')

def extract_tables_from_pdf(pdf_file_path):
    with open(pdf_file_path, 'rb') as file:
        pdf_bytes = file.read()

    # Request table extraction specifically
    response = textract.analyze_document(Document={'Bytes': pdf_bytes}, FeatureTypes=['TABLES'])

    # Parse JSON response to extract table data
    blocks = response['Blocks']
    tables = []
    for block in blocks:
        if block['BlockType'] == 'TABLE':
            rows = []
            for cell in block['Cells']:
                row_text = ""
                for relationship in cell['Relationships']:
                    if relationship['Type'] == 'CHILD':
                        row_text += get_text(relationship['Ids'], blocks)
                rows.append(row_text.split('\n'))
            tables.append(rows)

    return tables

def get_text(ids, blocks):
    text = ""
    for id in ids:
        for block in blocks:
            if block['Id'] == id:
                if block['BlockType'] == 'CELL':
                    text += block['Text'] + ' '
                elif block['BlockType'] == 'WORD':
                    text += block['Text'] + ' '
    return text

# Process multiple PDFs
for pdf_path in your_pdf_paths:
    extracted_tables = extract_tables_from_pdf(pdf_path)

    # Create Excel workbook and sheets for each table
    workbook = openpyxl.Workbook()
    for table_index, table in enumerate(extracted_tables):
        worksheet = workbook.create_sheet(f"Table_{table_index+1}")
        for row_index, row in enumerate(table):
            for col_index, cell_value in enumerate(row):
                worksheet.cell(row=row_index+1, column=col_index+1).value = cell_value

    # Save Excel file
    workbook.save(f"extracted_tables_{pdf_path.stem}.xlsx")

"""Key points:

Accuracy: Amazon Textract is often more accurate than local libraries for complex tables.
Processing Time: It can handle multiple PDFs quickly, utilizing AWS cloud resources.
Cost: Be mindful of costs associated with Amazon Textract usage.
"""